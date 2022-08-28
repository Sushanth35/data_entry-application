from logging import root
from msilib.schema import ComboBox
from operator import add
from optparse import Values
from sre_parse import State
from tkinter import*
from tkinter import ttk
import tkinter as tk
from tkinter import messagebox
from tkinter import font
from turtle import setx, width
from unicodedata import name
import openpyxl ,xlrd
from openpyxl import Workbook
import pathlib

root=Tk()
root.title("Data Entry")
root.geometry('700x400+300+200')
root.resizable(False,False)
root.configure(bg="#326273")

file=pathlib.Path('Backend_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Full Name"
    sheet['B1']="phone number"
    sheet['C1']="Age"
    sheet['D1']="Gender"
    sheet['E1']="Address"

    file.save('Backend_data.xlsx')



def submit():
    name=nameValue.get()
    contact= contactValue.get()
    age=AgeValue.get()
    sex=gender.get()
    address=addressEntry.get(1.0,END)

    file=openpyxl.load_workbook('Backend_Data.xlsx')
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=name)
    sheet.cell(column=2,row=sheet.max_row+1,value=contact)
    sheet.cell(column=3,row=sheet.max_row+1,value=age)
    sheet.cell(column=4,row=sheet.max_row+1,value=sex)
    sheet.cell(column=5,row=sheet.max_row+1,value=address)

    file.save('Backend_data.xlsx')









def clear():
    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    addressEntry.delete(1.0,END)





#heading
Label(root,text="please fill out this entry form:",font="arial 13",bg="#326273",fg="#fff").place(x=20,y=20)

#Label
Label(root,text="Name",font=23,bg="#326273",fg="#fff").place(x=50,y=100)
Label(root,text="Contact No.",font=23,bg="#326273",fg="#fff").place(x=50,y=150)
Label(root,text="Age",font=23,bg="#326273",fg="#fff").place(x=50,y=200)
Label(root,text="Gender",font=23,bg="#326273",fg="#fff").place(x=370,y=200)
Label(root,text="Address",font=23,bg="#326273",fg="#fff").place(x=50,y=250)

#Entry
nameValue = StringVar()
contactValue = StringVar()
AgeValue = StringVar()

nameEntry = Entry(root,textvariable=nameValue,width=45,bd=2,font=20)
contactEntry = Entry(root,textvariable=contactValue,width=45,bd=2,font=20)
ageEntry = Entry(root,textvariable=AgeValue,width=15,bd=2,font=20)


#gender
gender = tk.IntVar()
radiobutton_1 = tk.Radiobutton(root, text='Male', variable=gender, value=1)
radiobutton_1.pack()
radiobutton_2 = tk.Radiobutton(root, text='Female', variable=gender, value=2)
radiobutton_2.pack()


nameEntry.place(x=200,y=100)
contactEntry.place(x=200,y=150)
ageEntry.place(x=200,y=200)
radiobutton_1.place(x=440,y=200)
radiobutton_2.place(x=520,y=200)


#address
addressEntry = Text(root,width=50,height=4,bd=4)
addressEntry.place(x=200,y=250)

Button(root,text='submit',bg="#326273",fg='white',width=15,height=2,command=submit).place(x=200,y=350)
Button(root,text='clear',bg="#326273",fg='white',width=15,height=2,command=clear).place(x=340,y=350)
Button(root,text='Exit',bg="#326273",fg='white',width=15,height=2,command=root.destroy).place(x=480,y=350)


root.mainloop()